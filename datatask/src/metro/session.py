#!/usr/bin/python3
import json
import openpyxl
import requests
wb = openpyxl.load_workbook('/home/jiayi/xezing/script/devices.xlsx')
sh=wb[wb.sheetnames[0]]
print(sh.max_row)
sess = requests.Session()
header={"Content-Type": 'application/json'}
url_login='http://10.17.200.88/auth/v1/login'
form_login = {"username": "admin","password": "240be518fabd2724ddb6f04eeb1da5967448d7e831c08c8fa822809f74c720a9","captcha": "8888","clientType": "pc-web"}
resp=sess.post(url_login, data=json.dumps(form_login), headers=header)
print(resp.ok)
print(resp.json())
resp_json = json.loads(resp.content)
token=resp_json['data']['accessToken']

header={"Content-Type": 'application/json', 'jwt-token':token}

for row in sh.iter_rows(min_row=4,max_row=10,min_col=10,max_col=10,values_only=True):
    for cell in row:
        print(requests.get('http://10.17.200.88/groupdevice/v1/device/runinfo/%s' %(cell),headers=header).json())
