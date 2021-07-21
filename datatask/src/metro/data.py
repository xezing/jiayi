#!/usr/bin/python3
import openpyxl
import requests
#wb = openpyxl.load_workbook(r'D:\work\workspace\jiayi\datatask\files\设备中心-设备管理.xlsx')
wb = openpyxl.load_workbook('/home/jiayi/xezing/script/设备中心-设备管理.xlsx')
sh=wb[wb.sheetnames[0]]
print(sh.max_row)
headers={'jwt-token':'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJjbGllbnRUeXBlIjoicGMtd2ViIiwidG9rZW5JZCI6IjE0MTc0MDQyMjE5MjU0ODY1OTQiLCJsb2dpbk5hbWUiOiJhZG1pbiIsImV4cCI6MTYyNjg2NzY3NSwibGFuZyI6InpoX0NOIiwidG9rZW5UeXBlIjoiYWNjZXNzIiwiaWF0IjoxNjI2NzcwNDc1LCJ1c2VySWQiOiI3MDIzNGU0MmFmOTMxMWU2ODY1OWM1OThhNjdhNzFjNiJ9.gRegn5UqHPn8eVc7iaAxVL6VaJ6s7pixkA8IDT5YQMI'}

for row in sh.iter_rows(min_row=4,max_row=sh.max_row,min_col=10,max_col=10,values_only=True):
    for cell in row:
        print(cell)
        print(requests.get('http://10.17.200.88/groupdevice/v1/device/runinfo/%s' %(cell),headers).json())
        #print(requests.get('http://10.17.200.88/groupdevice/v1/device/runinfo/%s' % (cell)).json())