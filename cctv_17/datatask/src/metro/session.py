#!/usr/bin/python3
import codecs
import csv
import json
import openpyxl
import requests

wb = openpyxl.load_workbook('datatask/files/设备中心-设备管理.xlsx')
sh = wb[wb.sheetnames[0]]
print(sh.max_row)
sess = requests.Session()
header = {"Content-Type": 'application/json'}
url_login = 'http://10.17.200.88/auth/v1/login'
form_login = {"username": "admin", "password": "240be518fabd2724ddb6f04eeb1da5967448d7e831c08c8fa822809f74c720a9",
              "captcha": "8888", "clientType": "pc-web"}
resp = sess.post(url_login, data=json.dumps(form_login), headers=header)
print(resp.ok)
print(resp.json())
resp_json = json.loads(resp.content)
token = resp_json['data']['accessToken']
header = {"Content-Type": 'application/json', 'jwt-token': token}

# 文件部分
f = codecs.open(r'D:\workspace\jiayi\cctv_17\datatask\files\test.csv', 'w')
writer = csv.writer(f, lineterminator='\n')
writer.writerow(["设备IP", "接入协议", "设备名称", "设备ID", "获取结果"])

for row in sh.iter_rows(min_row=4, max_row=10, min_col=10, max_col=10, values_only=True):
    response_data = requests.get('http://10.17.200.88/groupdevice/v1/device/runinfo/%s' % (row[9]),
                                 headers=header).json()
    writer.writerow([row[2], row[8], row[1], row[9], response_data])
    for cell in row:
        print(requests.get('http://10.17.200.88/groupdevice/v1/device/runinfo/%s' % (cell), headers=header).json())
